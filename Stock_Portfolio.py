import openpyxl
import pandas as pd
import yfinance as yf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import math
from fuzzywuzzy import fuzz

def read_hold_file(hold_file_path):
    hold_df = pd.read_excel(hold_file_path)
    hold_df.columns = hold_df.columns.str.strip()
    return hold_df

def read_portfolio(portfolio_file, target_percentages_file):
    # Extract account information from the portfolio file
    account_info_df = pd.read_excel(portfolio_file, header=None, nrows=7)
    account_number = account_info_df.iloc[1, 1]
    account_name = account_info_df.iloc[2, 1]
    # Read the current portfolio data
    portfolio_df = pd.read_excel(portfolio_file, header=9)
    portfolio_df.columns = portfolio_df.columns.str.strip()
    # Clean and convert data types as necessary
    portfolio_df['Market Value'] = portfolio_df['Market Value'].replace({',': ''}, regex=True).astype(float)
    target_df = pd.read_csv(target_percentages_file)
    if '9999227' in portfolio_df['Symbol / CUSIP / ID'].values:
        cash_row = portfolio_df[portfolio_df['Symbol / CUSIP / ID'] == '9999227']
        previous_day_cash = cash_row['Market Value'].iloc[0]
    return portfolio_df, previous_day_cash, account_number, account_name

def is_account_match(account_name, account_number, hold_account_name, hold_account_number):
    # Check for 60% match in account name and exact match in account number
    name_match_score = fuzz.partial_ratio(account_name.lower(), hold_account_name.lower())
    return name_match_score >= 60 and account_number == hold_account_number

def find_matching_stocks(portfolio_df, hold_df, account_name, account_number):
    # Convert account numbers to string for both DataFrames
    hold_df['Account Number'] = hold_df['Account Number'].astype(str)
    account_number_str = str(account_number)

    # Filter hold_df for matching account number
    filtered_hold_df = hold_df[hold_df['Account Number'] == account_number_str]
    if len(filtered_hold_df) == 0:
        print("No rows found matching the account number. Sample data from hold_df:")
        print(hold_df.head())

    # Additional diagnostics for account name matching
    filtered_hold_df['Name Match Score'] = filtered_hold_df['Account'].apply(
        lambda x: fuzz.partial_ratio(account_name.lower(), x.lower()))
    filtered_hold_df = filtered_hold_df[filtered_hold_df['Name Match Score'] >= 60]

    # Find matching stock symbols
    matching_symbols = filtered_hold_df['Symbol'].isin(portfolio_df['Symbol / CUSIP / ID'])
    matching_stocks = filtered_hold_df[matching_symbols]
    non_matching_stocks = filtered_hold_df[~matching_symbols]

    # print("Details of matching stocks:")
    # print(matching_stocks[['Account Number', 'Account', 'Symbol', 'Quantity', 'Description']].to_string(index=False))
    # print("Details of non-matching stocks:")
    # print(non_matching_stocks[['Account Number', 'Account', 'Symbol', 'Quantity', 'Description']].to_string(index=False))

    return matching_stocks

def read_target_percentages(file_name):
    df = pd.read_csv(file_name)
    df['Target'] = df['Target'].astype(float)
    return df

def fetch_stock_prices(stock_list):
    stock_data = {}
    for stock in stock_list:
        if not isinstance(stock, str):
            continue
        # Set the price of '9999227' to 1
        if stock == '9999227':
            stock_data[stock] = {'price': 1, 'description': 'Insured Cash Account'}
            continue
        try:
            ticker = yf.Ticker(stock.upper())
            history = ticker.history(period="1d")
            if not history.empty:
                price = history['Close'].iloc[-1]
                description = ticker.info.get('longName', stock)
                stock_data[stock] = {'price': price, 'description': description}
                if stock == 'FDX':
                    price = 1
                    description = "Insured Cash Account"
                    stock_data[stock] = {"price": price, 'description': description}
            else:
                print(f"No trading data for {stock} on the requested date.")
                stock_data[stock] = {'price': None, 'description': None}
        except Exception as e:
            print(f"Error fetching data for ticker {stock}: {e}")
            stock_data[stock] = {'price': None, 'description': None}
    return stock_data

def adjust_quantity_based_on_hold(portfolio_df, matching_stocks, stock_data):
    portfolio_df['Current Quantity - Held Stocks'] = portfolio_df['Quantity']
    portfolio_df['Current Market Value - Held Stocks'] = 0  # Initialize with zero
    
    total_held_value = 0
    
    for index, hold_row in matching_stocks.iterrows():
        hold_symbol = hold_row['Symbol']
        hold_quantity = hold_row['Quantity']
        current_price = stock_data.get(hold_symbol, {}).get('price', 0)
        matching_portfolio_rows = portfolio_df[portfolio_df['Symbol / CUSIP / ID'] == hold_symbol]
        if not matching_portfolio_rows.empty:
            portfolio_index = matching_portfolio_rows.index[0]
            quantity_to_hold = min(hold_quantity, matching_portfolio_rows.at[portfolio_index, 'Quantity'])
            held_value = quantity_to_hold * current_price
            portfolio_df.at[portfolio_index, 'Current Quantity - Held Stocks'] -= quantity_to_hold
            portfolio_df.at[portfolio_index, 'Current Market Value - Held Stocks'] += held_value
            
            total_held_value += held_value
            
            # Print out the changes for each symbol
            print(f"Held {quantity_to_hold} of {hold_symbol}, Remaining in Portfolio: {portfolio_df.at[portfolio_index, 'Current Quantity - Held Stocks']}")

    print(f"Total value of held stocks: ${total_held_value:,.2f}")
    return portfolio_df, total_held_value

def calculate_required_shares(portfolio_df, target_df, stock_data, total_account_value, matching_stocks):
    adjusted_portfolio_df, total_held_value = adjust_quantity_based_on_hold(portfolio_df, matching_stocks, stock_data)
    print(f"Total account value before any allocation: {total_account_value}")
    print(f"Total value of held stocks: {total_held_value}\n")
    remaining_account_value = total_account_value - total_held_value

    # Prepare the data for the current portfolio status before any stocks are held
    current_results = []
    for index, row in portfolio_df.iterrows():
        symbol = row['Symbol / CUSIP / ID']
        recent_price_info = stock_data.get(symbol)
        
        if recent_price_info and recent_price_info['price'] is not None:
            recent_price = recent_price_info['price']
            original_quantity = row['Quantity']
            held_quantity = adjusted_portfolio_df.at[index, 'Current Quantity - Held Stocks']
            held_quantity = original_quantity - held_quantity
            original_market_value = original_quantity * recent_price
            held_market_value = held_quantity * recent_price
            
            # Append the original and held stock information to the results
            current_results.append({
                'Current Symbol': symbol,
                'Current Description': row['Description / Fund'],
                'Current Quantity': original_quantity,
                'Current Price of Stock': "${:,.2f}".format(recent_price),
                'Current Market Value': "${:,.2f}".format(original_market_value),
                'Held Quantity': held_quantity,
                'Held Market Value': "${:,.2f}".format(held_market_value),
                'Current Percent of Holding': "{:.2f}%".format(row['Percent of Account Holdings'] * 100)
            })
        else:
            print(f"No recent price data for {symbol}")

    # DataFrame for current results
    current_results_df = pd.DataFrame(current_results)
    target_results = []
    remaining_funds = remaining_account_value

    for _, target_row in target_df.iterrows():
        target_symbol = target_row['Security Ticker/CUSIP']
        target_percentage = target_row['Target'] / 100
        recent_price_info = stock_data.get(target_symbol)
        
        if recent_price_info and recent_price_info['price'] is not None:
            recent_price = recent_price_info['price']
            target_allocation = remaining_account_value * target_percentage
            target_quantity = math.floor(target_allocation / recent_price)
            target_market_value = target_quantity * recent_price
            remaining_funds -= target_market_value

            target_results.append({
                'Target Symbol': target_symbol,
                'Target Description': recent_price_info['description'],
                'Target Quantity': target_quantity,
                'Target Price of Stock': "${:,.2f}".format(recent_price),
                'Target Market Value': target_market_value
            })
        else:
            print(f"No recent price data for {target_symbol}")

    # Now, format 'Target Market Value' as a currency string and calculate 'Target Percent of Holding'
    for item in target_results:
        item['Target Percent of Holding'] = "{:.2f}%".format((item['Target Market Value'] / total_account_value) * 100)
        item['Target Market Value'] = "${:,.2f}".format(item['Target Market Value'])

    leverage_stock = 0.01
    leverage_fdx = 0.0075
    for target_result in target_results:
        if remaining_funds <= 0:
            break

        target_symbol = target_result['Target Symbol']
        target_percentage = float(target_result['Target Percent of Holding'].strip('%')) / 100
        recent_price_info = stock_data.get(target_symbol)

        if recent_price_info and recent_price_info['price'] is not None:
            recent_price = recent_price_info['price']
            leverage = leverage_stock if target_symbol != 'FDX' else leverage_fdx
            # Convert 'Target Market Value' to a float for arithmetic operations
            target_market_value = float(target_result['Target Market Value'].replace('$', '').replace(',', ''))
            max_additional_allocation = total_account_value * (target_percentage + leverage) - target_market_value
            additional_quantity = min(math.floor(remaining_funds / recent_price), math.floor(max_additional_allocation / recent_price))
            
            if additional_quantity > 0:
                additional_market_value = additional_quantity * recent_price
                remaining_funds -= additional_market_value
                # Convert 'Target Market Value' from string to float, add the value, then convert back to string if necessary
                target_market_value_float = float(target_result['Target Market Value'].replace('$', '').replace(',', ''))
                target_market_value_float += additional_market_value
                target_result['Target Quantity'] += round(additional_quantity, 2)
                target_result['Target Percent of Holding'] = "{:.2f}%".format((target_market_value_float / remaining_account_value) * 100)
                target_result['Target Market Value'] = "${:,.2f}".format(target_market_value_float)

    # Allocate remaining funds to FDX stock (cash account), adjusting for two decimal places
    if remaining_funds > 0:
        fdx_price = float(stock_data.get('FDX', {}).get('price', 1))
        max_fdx_allocation = total_account_value * 0.015  # 1.5% of the total account value
        fdx_entry = next((item for item in target_results if item['Target Symbol'] == 'FDX'), None)
        if fdx_entry:
            # Convert 'Target Market Value' from string to float for arithmetic operation
            fdx_market_value_float = float(fdx_entry['Target Market Value'].replace('$', '').replace(',', ''))
            fdx_price = stock_data['FDX']['price']
            additional_fdx_market_value = min(remaining_funds, max_fdx_allocation - fdx_market_value_float)
            additional_fdx_quantity = additional_fdx_market_value / fdx_price
            fdx_entry['Target Quantity'] += round(additional_fdx_quantity, 2)
            fdx_market_value_float += additional_fdx_market_value
            fdx_entry['Target Market Value'] = "${:,.2f}".format(fdx_market_value_float)
            fdx_entry['Target Percent of Holding'] = (fdx_market_value_float / remaining_account_value) * 100
        else:
            print("  No additional FDX allocation needed or FDX not found.")
            additional_fdx_market_value = min(remaining_funds, max_fdx_allocation)
            additional_fdx_quantity = additional_fdx_market_value / fdx_price
            target_results.append({
                'Target Symbol': 'FDX',
                'Target Description': 'Insured Cash Account',
                'Target Quantity': round(additional_fdx_quantity, 2),
                'Target Price of Stock': "${:,.2f}".format(fdx_price),
                'Target Market Value': "${:,.2f}".format(additional_fdx_market_value),
                'Target Percent of Holding': "{:.2f}%".format((additional_fdx_market_value / remaining_account_value) * 100)
            })    
    total_target_market_value = sum(float(item['Target Market Value'].replace('$', '').replace(',', '')) for item in target_results)

    for item in target_results:
        # Format the 'Target Market Value' as a numeric type for the final percentage calculation
        item['Target Market Value'] = float(item['Target Market Value'].replace('$', '').replace(',', ''))
        # Calculate 'Target Percent of Holding' as a percentage of the total account value, after all allocations
        item['Target Percent of Holding'] = (item['Target Market Value'] / total_target_market_value) * 100
        # Then, format 'Target Market Value' back as a currency string
        item['Target Market Value'] = "${:,.2f}".format(item['Target Market Value'])
        # Format 'Target Percent of Holding' as a string with percentage format
        item['Target Percent of Holding'] = "{:.2f}%".format(item['Target Percent of Holding'])

    target_results_df = pd.DataFrame(target_results)
    results_df = pd.concat([current_results_df, target_results_df], axis=1)
    results_df = results_df[[column for column in results_df.columns if 'Target Percent Holdings' not in column]]
    return results_df

def add_summary_rows(df):
    # Convert percentage strings to floats if needed and sum
    if df['Current Percent of Holding'].dtype == object:
        df['Current Percent of Holding'] = df['Current Percent of Holding'].str.replace('%', '').astype(float)
    else:
        df['Current Percent of Holding'] = df['Current Percent of Holding'].astype(float)

    # Sum up the total current percent
    current_percent_total = df['Current Percent of Holding'].sum()
    
    if df['Target Percent of Holding'].dtype == object:
        df['Target Percent of Holding'] = df['Target Percent of Holding'].str.replace('%', '').astype(float)
    else:
        df['Target Percent of Holding'] = df['Target Percent of Holding'].astype(float)

    # Sum up the total target percent
    target_percent_total = df['Target Percent of Holding'].sum()
    
    # Remove '$' and ',' from 'Current Market Value' and 'Target Market Value', convert to float, and sum
    df['Current Market Value'] = df['Current Market Value'].replace('[\$,]', '', regex=True).astype(float)
    current_market_value_total = df['Current Market Value'].sum()

    df['Target Market Value'] = df['Target Market Value'].replace('[\$,]', '', regex=True).astype(float)
    target_market_value_total = df['Target Market Value'].sum()

    # For 'Current Price of Stock' and 'Target Price of Stock', remove '$' and ',' then convert to float for sum
    df['Current Price of Stock'] = df['Current Price of Stock'].replace('[\$,]', '', regex=True).astype(float)
    current_price_total = df['Current Price of Stock'].sum()
    
    df['Held Market Value'] = df['Held Market Value'].replace('[\$,]', '', regex=True).astype(float)
    held_market_value_total = df['Held Market Value'].sum()

    df['Target Price of Stock'] = df['Target Price of Stock'].replace('[\$,]', '', regex=True).astype(float)
    target_price_total = df['Target Price of Stock'].sum()

    # Format the summary values
    formatted_summary = {
        'Current Symbol': "Summary",
        'Current Description': "",
        'Current Quantity': df['Current Quantity'].sum(),
        'Current Price of Stock': "${:,.2f}".format(current_price_total),
        'Current Market Value': "${:,.2f}".format(current_market_value_total),
        'Held Quantity' : df['Held Quantity'].astype(float).sum(),
        'Held Market Value': "{:,.2f}".format(held_market_value_total),
        'Current Percent of Holding': "{:.2f}%".format(current_percent_total),
        'Target Symbol': "",
        'Target Description': "",
        'Target Quantity': df['Target Quantity'].sum(),
        'Target Price of Stock': "${:,.2f}".format(target_price_total),
        'Target Market Value': "${:,.2f}".format(target_market_value_total),
        'Target Percent of Holding': "{:.2f}%".format(target_percent_total)
    }

    # Convert values back to $ or % and append the summary row to the DataFrame
    df['Current Percent of Holding'] = df['Current Percent of Holding'].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else x)
    df['Target Percent of Holding'] = df['Target Percent of Holding'].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else x)
    df['Current Market Value'] = df['Current Market Value'].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else x)
    df['Target Market Value'] = df['Target Market Value'].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else x)
    df['Held Market Value'] = df['Held Market Value'].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else x)
    df['Current Price of Stock'] = df['Current Price of Stock'].apply(lambda x: f"${x:.2f}" if pd.notnull(x) else x)
    df['Target Price of Stock'] = df['Target Price of Stock'].apply(lambda x: f"${x:.2f}" if pd.notnull(x) else x)

    # Replace NaN values with empty strings
    df_with_summary = pd.concat([df, pd.DataFrame([formatted_summary])], ignore_index=True)
    return df_with_summary

def write_to_excel(client_name, account_number, data_frame, filename):
    wb = Workbook()
    ws = wb.active

    # Set the header with larger font, white text, and blue background
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=18)
    title = f"{client_name} - Account {account_number}"
    ws.append([title] + [''] * (len(data_frame.columns) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data_frame.columns))
    for cell in ws["1:1"]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = [
        'Current Symbol', 'Current Description', 'Current Quantity', 'Current Price of Stock', 'Current Market Value','Held Quantity', 'Held Market Value', 'Current Percent of Holding', 
        'Target Symbol', 'Target Description','Target Quantity', 'Target Price of Stock', 'Target Market Value', 'Target Percent of Holding'
    ]
    ws.append(headers)
    
    # Data
    for r_idx, row in enumerate(dataframe_to_rows(data_frame, index=False, header=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    format_excel_sheet(ws)
    wb.save(filename)
    
def as_text(value):
    return str(value) if value is not None else ""

def auto_adjust_column_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length


def format_excel_sheet(ws):
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    font_white = Font(color="FFFFFF", bold=True, size=12)
    font_regular = Font(size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Apply styles to header row
    for cell in ws['1']:
        cell.fill = header_fill
        cell.font = font_white
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='thick'))
        cell.font = Font(color = "FFFFFF", size=16, bold = True)

    # Apply styles to data cells
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row - 1):
        for cell in row:
            cell.font = font_regular
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')

            # Apply alternate background color for easier reading
            if cell.row % 2 == 0:
                cell.fill = alt_fill

    # Style the summary row
    summary_row = ws[ws.max_row]
    for cell in summary_row:
        cell.fill = header_fill
        cell.font = font_white
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
    auto_adjust_column_width(ws)

def create_trading_upload_output(account_number, account_name, final_df):
    columns_required = ['Target Symbol', 'Target Quantity', 'Target Market Value', 'Target Description']
    trading_upload_df = final_df[columns_required].copy()
    trading_upload_df.columns = ['Symbol', 'Quantity', 'Allocation ($)', 'Description']
    trading_upload_df.insert(0, 'Account', account_name)
    trading_upload_df.insert(0, 'Account Number', account_number)
    trading_upload_df = trading_upload_df[trading_upload_df['Symbol'].notna()]
    return trading_upload_df

def write_trading_upload_to_excel(account_name, trading_upload_df):
    current_date_time = datetime.now().strftime('%m_%d_%Y_%H%M%S')
    output_filename = f"Trading Upload_{account_name}_{current_date_time}.xlsx"
    trading_upload_df.to_excel(output_filename, index=False)
    print(f"File saved: {output_filename}")


def main():
    portfolio_file = 'Investments-Summary-12-1-2023.xlsx'
    target_percentages_file = 'SampleSecurityImport_EQ ARG GR id 77.csv'
    hold_file = 'Client_HoldStocks_12-1-2023.xlsx'
    
    # Read the portfolio and hold files
    portfolio_df, previous_day_cash, account_number, account_name = read_portfolio(portfolio_file, target_percentages_file)
    hold_df = read_hold_file(hold_file)

    # Get matching stocks
    matching_stocks = find_matching_stocks(portfolio_df, hold_df, account_name, account_number)

    # Print or write the matching stocks
    if not matching_stocks.empty:
        print("Matching stocks found:")
        print(matching_stocks[['Account Number', 'Account', 'Symbol', 'Quantity', 'Description']].to_string(index=False))
    else:
        print("No matching stocks found for the matching account.")
        
    target_df = read_target_percentages(target_percentages_file)
    if target_df is None:
        raise ValueError("target_df is None. Check the read_target_percentages function.")
    if 'Target' not in target_df.columns:
        raise ValueError("Column 'Target' not found in target DataFrame.")
    if 'Symbol / CUSIP / ID' not in portfolio_df.columns:
        raise ValueError("Column 'Symbol / CUSIP / ID' not found in portfolio DataFrame.")
    # Fetch stock prices and details
    all_stocks = list(set(portfolio_df['Symbol / CUSIP / ID']) | set(target_df['Security Ticker/CUSIP']))
    stock_data = fetch_stock_prices(all_stocks)
    total_account_value = 0
    for stock in portfolio_df['Symbol / CUSIP / ID']:
        stock_row = portfolio_df[portfolio_df['Symbol / CUSIP / ID'] == stock]
        if not stock_row.empty and stock in stock_data and 'price' in stock_data[stock]:
            total_account_value += stock_data[stock]['price'] * stock_row['Quantity'].iloc[0]
    matching_stocks = find_matching_stocks(portfolio_df, hold_df, account_name, account_number)
    results_df = calculate_required_shares(portfolio_df, target_df, stock_data, total_account_value, matching_stocks)
    final_df = add_summary_rows(results_df)
    trading_upload_df = create_trading_upload_output(account_number, account_name, final_df)
    current_date = datetime.now().strftime('%m_%d_%Y')
    output_filename = f"{account_name}_SharesOfStocks_{current_date}.xlsx"
    # Write to Excel using the updated account information
    write_to_excel(account_name, account_number, final_df, output_filename)
    write_trading_upload_to_excel(account_name, trading_upload_df)

if __name__ == "__main__":
    main()